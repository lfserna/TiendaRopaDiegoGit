import os
import random
import re
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional

from flask import current_app
from openpyxl import Workbook, load_workbook


class ExcelService:
    tz_utc_minus_4 = timezone(timedelta(hours=-4))

    pedidos_columns = [
        "id",
        "pedido_grupo_id",
        "codigo_producto",
        "codigo_principal",
        "registro_principal",
        "es_copia",
        "nombre",
        "apellido",
        "telefono",
        "tipo_entrega",
        "ciudad_destino",
        "lugar_entrega",
        "fecha_entrega",
        "hora_entrega",
        "direccion_referencia",
        "costo_total_pedido",
        "estado",
        "fecha_registro",
        "fecha_actualizacion",
        "observaciones",
    ]

    codigos_columns = [
        "id",
        "codigo",
        "estado_codigo",
        "fecha_creacion",
        "fecha_uso",
        "pedido_grupo_id",
        "nombre",
        "telefono",
        "generado_por",
        "observaciones",
    ]

    config_sheets = {
        "ciudades": ["La Paz", "Oruro", "Potosi", "Cochabamba", "Tarija", "Chuquisaca", "Pando", "Beni", "Montero"],
        "lugares": ["Centro", "Plan 3000", "Cambodromo", "Equipetrol"],
        "horarios": ["09:00", "11:00", "15:00", "18:00"],
        "dias": ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado"],
        "estados": [
            "Pendiente",
            "Confirmado",
            "En preparacion",
            "Listo para entregar",
            "Enviado",
            "Entregado",
            "No entregado",
            "Cancelado",
        ],
    }

    def ensure_files(self) -> None:
        os.makedirs(current_app.config["REGISTROS_DIR"], exist_ok=True)
        os.makedirs(current_app.config["CONFIG_DIR"], exist_ok=True)

        if not os.path.exists(current_app.config["PEDIDOS_FILE"]):
            self._create_pedidos_file()

        if not os.path.exists(current_app.config["CODIGOS_FILE"]):
            self._create_codigos_file()

        if not os.path.exists(current_app.config["CONFIG_FILE"]):
            self._create_config_file()

        self._normalize_numeric_ids(current_app.config["PEDIDOS_FILE"], "pedidos")
        self._normalize_numeric_ids(current_app.config["CODIGOS_FILE"], "codigos")
        self._normalize_config_ids()

    def _now_str(self) -> str:
        return datetime.now(self.tz_utc_minus_4).strftime("%Y-%m-%d / %H:%M:%S")

    def _next_group_id(self) -> str:
        orders = self.read_orders()
        pattern = re.compile(r"^PG-(\d{5})$")
        last_number = 0

        for order in orders:
            group_id = str(order.get("pedido_grupo_id") or "").strip()
            match = pattern.match(group_id)
            if match:
                last_number = max(last_number, int(match.group(1)))

        return f"PG-{last_number + 1:05d}"

    def _create_pedidos_file(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "pedidos"
        ws.append(self.pedidos_columns)
        wb.save(current_app.config["PEDIDOS_FILE"])

    def _create_codigos_file(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "codigos"
        ws.append(self.codigos_columns)
        wb.save(current_app.config["CODIGOS_FILE"])

    def _normalize_numeric_ids(self, file_path: str, sheet_name: str) -> None:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            row[0].value = idx

        wb.save(file_path)
        wb.close()

    def _next_numeric_id(self, file_path: str, sheet_name: str) -> int:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        values = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                try:
                    values.append(int(row[0]))
                except (TypeError, ValueError):
                    continue

        wb.close()
        return (max(values) + 1) if values else 1

    def _next_config_sheet_id(self, sheet_name: str) -> int:
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]
        values = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                try:
                    values.append(int(row[0]))
                except (TypeError, ValueError):
                    continue

        wb.close()
        return (max(values) + 1) if values else 1

    def normalize_code(self, code: str) -> str:
        raw = "".join(ch for ch in (code or "").upper() if ch.isalnum())
        if not raw:
            return ""

        parts = []
        index = 0
        for length in (2, 2, 1):
            segment = raw[index:index + length]
            if not segment:
                break
            parts.append(segment)
            index += len(segment)

        return "-".join(parts)

    def is_valid_code_format(self, code: str) -> bool:
        raw = "".join(ch for ch in (code or "").upper() if ch.isalnum())
        if len(raw) != 5:
            return False

        normalized = self.normalize_code(code)
        if len(normalized) != 7:
            return False
        if normalized.count("-") != 2:
            return False
        left, middle, right = normalized.split("-")
        return (
            len(left) == 2
            and len(middle) == 2
            and len(right) == 1
            and left.isalpha()
            and middle.isdigit()
            and right.isalpha()
        )

    def _create_config_file(self) -> None:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, options in self.config_sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["id", "valor", "activo", "orden", "fecha_actualizacion"])
            for idx, value in enumerate(options, start=1):
                ws.append([idx, value, True, idx, self._now_str()])

        wb.save(current_app.config["CONFIG_FILE"])

    def _normalize_config_ids(self) -> None:
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        for sheet_name in self.config_sheets.keys():
            ws = wb[sheet_name]
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                row[0].value = idx
        wb.save(current_app.config["CONFIG_FILE"])
        wb.close()

    def _read_sheet_rows(self, sheet_name: str) -> List[Dict]:
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})

        wb.close()
        rows.sort(key=lambda x: (x.get("orden") or 9999, (x.get("valor") or "")))
        return rows

    def get_config_options(self, include_inactive: bool = False) -> Dict[str, List[Dict]]:
        config_data = {}
        for sheet in self.config_sheets.keys():
            rows = self._read_sheet_rows(sheet)
            if not include_inactive:
                rows = [r for r in rows if bool(r.get("activo"))]
            config_data[sheet] = rows
        return config_data

    def add_config_item(self, sheet_name: str, value: str) -> None:
        value = value.strip()
        if not value:
            raise ValueError("El valor no puede estar vacio")

        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and str(row[1]).lower() == value.lower():
                wb.close()
                raise ValueError("Ese valor ya existe")

        next_order = ws.max_row
        next_id = self._next_config_sheet_id(sheet_name)
        ws.append([next_id, value, True, next_order, self._now_str()])
        wb.save(current_app.config["CONFIG_FILE"])
        wb.close()

    def update_config_item(self, sheet_name: str, item_id: str, new_value: str) -> None:
        new_value = new_value.strip()
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == item_id:
                row[1].value = new_value
                row[4].value = self._now_str()
                wb.save(current_app.config["CONFIG_FILE"])
                wb.close()
                return

        wb.close()
        raise ValueError("No se encontro el item")

    def toggle_config_item(self, sheet_name: str, item_id: str, active: bool) -> None:
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == item_id:
                row[2].value = active
                row[4].value = self._now_str()
                wb.save(current_app.config["CONFIG_FILE"])
                wb.close()
                return

        wb.close()
        raise ValueError("No se encontro el item")

    def delete_config_item(self, sheet_name: str, item_id: str) -> None:
        wb = load_workbook(current_app.config["CONFIG_FILE"])
        ws = wb[sheet_name]

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == item_id:
                ws.delete_rows(idx, 1)
                wb.save(current_app.config["CONFIG_FILE"])
                wb.close()
                return

        wb.close()
        raise ValueError("No se encontro el item")

    def read_orders(self) -> List[Dict]:
        wb = load_workbook(current_app.config["PEDIDOS_FILE"])
        ws = wb["pedidos"]
        headers = [cell.value for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})

        wb.close()
        return rows

    def read_codes(self) -> List[Dict]:
        wb = load_workbook(current_app.config["CODIGOS_FILE"])
        ws = wb["codigos"]
        headers = [cell.value for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})

        wb.close()
        return rows

    def _build_random_code(self) -> str:
        letters = "ABCDEFGHJKLMNPQRSTUVWXYZ"
        left = "".join(random.choice(letters) for _ in range(2))
        middle = f"{random.randint(0, 99):02d}"
        right = random.choice(letters)
        return f"{left}-{middle}-{right}"

    def generate_next_code(self, generado_por: str = "admin") -> Dict:
        existing = {str(item.get("codigo", "")).upper() for item in self.read_codes()}

        code = self._build_random_code()
        while code.upper() in existing:
            code = self._build_random_code()

        now = self._now_str()
        next_id = self._next_numeric_id(current_app.config["CODIGOS_FILE"], "codigos")
        row = {
            "id": next_id,
            "codigo": code,
            "estado_codigo": "Disponible",
            "fecha_creacion": now,
            "fecha_uso": "",
            "pedido_grupo_id": "",
            "nombre": "",
            "telefono": "",
            "generado_por": generado_por,
            "observaciones": "",
        }

        wb = load_workbook(current_app.config["CODIGOS_FILE"])
        ws = wb["codigos"]
        ws.append([row[col] for col in self.codigos_columns])
        wb.save(current_app.config["CODIGOS_FILE"])
        wb.close()

        return row

    def get_code_record(self, code: str) -> Optional[Dict]:
        normalized = (code or "").strip().upper()
        if not normalized:
            return None

        for item in self.read_codes():
            if str(item.get("codigo", "")).strip().upper() == normalized:
                return item
        return None

    def is_code_available(self, code: str) -> bool:
        record = self.get_code_record(code)
        if not record:
            return False
        return str(record.get("estado_codigo", "")).strip().lower() == "disponible"

    def mark_code_as_used(self, code: str, pedido_grupo_id: str, nombre: str, telefono: str) -> bool:
        normalized = (code or "").strip().upper()
        wb = load_workbook(current_app.config["CODIGOS_FILE"])
        ws = wb["codigos"]
        updated = False

        for row in ws.iter_rows(min_row=2):
            row_code = str(row[1].value or "").strip().upper()
            if row_code == normalized:
                row[2].value = "Usado"
                row[4].value = self._now_str()
                row[5].value = pedido_grupo_id
                row[6].value = nombre
                row[7].value = telefono
                updated = True
                break

        if updated:
            wb.save(current_app.config["CODIGOS_FILE"])

        wb.close()
        return updated

    def get_recent_codes(self, limit: int = 20) -> List[Dict]:
        codes = self.read_codes()
        codes.sort(key=lambda x: x.get("fecha_creacion") or "", reverse=True)
        return codes[:limit]

    def find_orders_by_code(self, code: str) -> List[Dict]:
        normalized = code.strip().upper()
        orders = self.read_orders()
        return [
            o
            for o in orders
            if str(o.get("codigo_producto", "")).upper() == normalized
            or str(o.get("codigo_principal", "")).upper() == normalized
        ]

    def get_related_orders(self, order: Dict) -> List[Dict]:
        orders = self.read_orders()

        if order.get("pedido_grupo_id"):
            group_id = order["pedido_grupo_id"]
            related = [o for o in orders if o.get("pedido_grupo_id") == group_id]
            if related:
                return sorted(related, key=lambda x: (x.get("registro_principal") is not True, x.get("fecha_registro") or ""))

        name = str(order.get("nombre", "")).strip().lower()
        phone = str(order.get("telefono", "")).strip()
        related = [
            o
            for o in orders
            if str(o.get("nombre", "")).strip().lower() == name
            and str(o.get("telefono", "")).strip() == phone
        ]
        return sorted(related, key=lambda x: (x.get("registro_principal") is not True, x.get("fecha_registro") or ""))

    def create_order_group(self, payload: Dict) -> str:
        main_code = self.normalize_code(payload["codigo_principal"])
        extra_codes = [self.normalize_code(c) for c in payload.get("codigos_extra", []) if c and c.strip()]

        unique_codes = [main_code]
        for code in extra_codes:
            if code not in unique_codes:
                unique_codes.append(code)

        group_id = self._next_group_id()
        now = self._now_str()

        wb = load_workbook(current_app.config["PEDIDOS_FILE"])
        ws = wb["pedidos"]
        next_id = self._next_numeric_id(current_app.config["PEDIDOS_FILE"], "pedidos")

        for idx, code in enumerate(unique_codes):
            is_main = idx == 0
            ws.append([
            next_id,
                group_id,
                code,
                main_code,
                is_main,
                not is_main,
                payload.get("nombre"),
                payload.get("apellido", ""),
                payload.get("telefono"),
                payload.get("tipo_entrega"),
                payload.get("ciudad_destino", ""),
                payload.get("lugar_entrega", ""),
                payload.get("fecha_entrega", ""),
                payload.get("hora_entrega", ""),
                payload.get("direccion_referencia", ""),
                payload.get("costo_total_pedido") if is_main else "",
                "Pendiente",
                now,
                now,
                payload.get("observaciones", ""),
            ])
            next_id += 1

        wb.save(current_app.config["PEDIDOS_FILE"])
        wb.close()

        return group_id

    def update_order_status(self, order_id: str, new_status: str) -> bool:
        wb = load_workbook(current_app.config["PEDIDOS_FILE"])
        ws = wb["pedidos"]
        updated = False

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(order_id):
                row[16].value = new_status
                row[18].value = self._now_str()
                updated = True
                break

        if updated:
            wb.save(current_app.config["PEDIDOS_FILE"])

        wb.close()
        return updated

    def update_order_status_by_code(self, code: str, new_status: str) -> bool:
        normalized = self.normalize_code(code)
        wb = load_workbook(current_app.config["PEDIDOS_FILE"])
        ws = wb["pedidos"]
        updated = False

        for row in ws.iter_rows(min_row=2):
            row_code = self.normalize_code(str(row[2].value or ""))
            if row_code == normalized:
                row[16].value = new_status
                row[18].value = self._now_str()
                updated = True
                break

        if updated:
            wb.save(current_app.config["PEDIDOS_FILE"])

        wb.close()
        return updated


excel_service = ExcelService()
